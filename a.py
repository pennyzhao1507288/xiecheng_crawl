import random
import time
import requests
from bs4 import BeautifulSoup
import json
from openpyxl import Workbook
import os
# 获取当前运行文件的所在目录
from get_hotel_id import save_hotel_ids_to_txt

current_directory = os.path.dirname(os.path.abspath(__file__))

"""
def get_comments(hotelId,pageIndex):
    burp0_url = "https://m.ctrip.com:443/restapi/soa2/24626/commentlist?_fxpcqlniredt=52271115296426477628&x-traceID=52271115296426477628-1697011438794-6881577"
    burp0_cookies = {"GUID": "52271115296426477628", "nfes_isSupportWebP": "1", "nfes_isSupportWebP": "1", "UBT_VID": "1697008100680.1e1cmmiBfNEU", "librauuid": "", "_RF1": "111.22.74.62", "_RSG": "2Ha6Cu221L4XAzI9a6MT2B", "_RDG": "282a01bdfbf8df2e6720e1b8de10ac1c86", "_RGUID": "4541082e-0c1a-4442-b09d-5869094bd2a1", "login_type": "0", "login_uid": "B34F49EE0A13E8463539186D23B00A8B", "DUID": "u=B34F49EE0A13E8463539186D23B00A8B&v=0", "IsNonUser": "F", "AHeadUserInfo": "VipGrade=10&VipGradeName=%BB%C6%BD%F0%B9%F3%B1%F6&UserName=&NoReadMessageCount=0", "cticket": "8A38C2D39415D5002658D2670FDD4F8D369AEE5E25A10B5F129505F540395603", "_pd": "%7B%22_o%22%3A3%2C%22s%22%3A130%2C%22_s%22%3A0%7D", "_bfa": "1.1697008100680.1e1cmmiBfNEU.1.1697010109935.1697011438385.1.19.228032", "Union": "OUID=&AllianceID=262684&SID=711465&SourceID=55552689&createtime=1697011439&Expires=1697616238511", "MKT_OrderClick": "ASID=262684711465&AID=262684&CSID=711465&OUID=&CT=1697011438513&CURL=https%3A%2F%2Fm.ctrip.com%2Fwebapp%2Fhotels%2Fcommentlist%3Fid%3D1632483%26hotelname%3D%25E6%2598%259F%25E8%25BE%25B0%25E9%2585%2592%25E5%25BA%2597(%25E6%25A0%25AA%25E6%25B4%25B2%25E6%25B9%2596%25E5%258D%2597%25E5%25B7%25A5%25E4%25B8%259A%25E5%25A4%25A7%25E5%25AD%25A6%25E5%25BA%2597)%26biz%3D1%26ftype%3Dv%26fromminiapp%3Dweixin%26allianceid%3D262684%26sid%3D711465%26sourceid%3D55552689%26_cwxobj%3D%257B%2522cid%2522%253A%252252271115296426477628%2522%252C%2522appid%2522%253A%2522wx0e6ed4f51db9d078%2522%252C%2522mpopenid%2522%253A%2522b9787617-697b-4fe6-9942-f5fddd4b8063%2522%252C%2522mpunionid%2522%253A%2522oHkqHt8Zg6-uBtF-Y2UBUVVk1MlM%2522%252C%2522allianceid%2522%253A%2522262684%2522%252C%2522sid%2522%253A%2522711465%2522%252C%2522ouid%2522%253A%2522%2522%252C%2522sourceid%2522%253A%252255552689%2522%252C%2522exmktID%2522%253A%2522%257B%255C%2522openid%255C%2522%253A%255C%2522b9787617-697b-4fe6-9942-f5fddd4b8063%255C%2522%252C%255C%2522unionid%255C%2522%253A%255C%2522oHkqHt8Zg6-uBtF-Y2UBUVVk1MlM%255C%2522%252C%255C%2522channelUpdateTime%255C%2522%253A%255C%25221697011325536%255C%2522%252C%255C%2522serverFrom%255C%2522%253A%255C%2522WAP%252FWECHATAPP%255C%2522%252C%255C%2522innersid%255C%2522%253A%255C%2522%255C%2522%252C%255C%2522innerouid%255C%2522%253A%255C%2522%255C%2522%252C%255C%2522pushcode%255C%2522%253A%255C%2522%255C%2522%252C%255C%2522txCpsId%255C%2522%253A%255C%2522%255C%2522%252C%255C%2522amsPid%255C%2522%253A%255C%2522%255C%2522%252C%255C%2522gdt_vid%255C%2522%253A%255C%2522%255C%2522%257D%2522%252C%2522scene%2522%253A1007%252C%2522personalRecommendSwitch%2522%253Atrue%252C%2522localRecommendSwitch%2522%253Atrue%252C%2522marketSwitch%2522%253Atrue%252C%2522pLen%2522%253A3%257D%26_obt%3D1697011436515&VAL={\"pc_vid\":\"1697008100680.1e1cmmiBfNEU\"}", "hotelhst": "1164390341"}
    burp0_headers = {"Sec-Ch-Ua": "\"Chromium\";v=\"107\", \"Not=A?Brand\";v=\"24\"", "Cookieorigin": "https://m.ctrip.com", "Content-Type": "application/json", "Sec-Ch-Ua-Mobile": "?0", "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/107.0.0.0 Safari/537.36 MicroMessenger/7.0.20.1781(0x6700143B) NetType/WIFI MiniProgramEnv/Windows WindowsWechat/WMPF XWEB/8447", "Sec-Ch-Ua-Platform": "\"Windows\"", "Accept": "*/*", "Origin": "https://m.ctrip.com", "Sec-Fetch-Site": "same-origin", "Sec-Fetch-Mode": "cors", "Sec-Fetch-Dest": "empty", "Referer": "https://m.ctrip.com/webapp/hotels/commentlist?id=1632483&hotelname=%E6%98%9F%E8%BE%B0%E9%85%92%E5%BA%97(%E6%A0%AA%E6%B4%B2%E6%B9%96%E5%8D%97%E5%B7%A5%E4%B8%9A%E5%A4%A7%E5%AD%A6%E5%BA%97)&biz=1&ftype=v&fromminiapp=weixin&allianceid=262684&sid=711465&sourceid=55552689&_cwxobj=%7B%22cid%22%3A%2252271115296426477628%22%2C%22appid%22%3A%22wx0e6ed4f51db9d078%22%2C%22mpopenid%22%3A%22b9787617-697b-4fe6-9942-f5fddd4b8063%22%2C%22mpunionid%22%3A%22oHkqHt8Zg6-uBtF-Y2UBUVVk1MlM%22%2C%22allianceid%22%3A%22262684%22%2C%22sid%22%3A%22711465%22%2C%22ouid%22%3A%22%22%2C%22sourceid%22%3A%2255552689%22%2C%22exmktID%22%3A%22%7B%5C%22openid%5C%22%3A%5C%22b9787617-697b-4fe6-9942-f5fddd4b8063%5C%22%2C%5C%22unionid%5C%22%3A%5C%22oHkqHt8Zg6-uBtF-Y2UBUVVk1MlM%5C%22%2C%5C%22channelUpdateTime%5C%22%3A%5C%221697011325536%5C%22%2C%5C%22serverFrom%5C%22%3A%5C%22WAP%2FWECHATAPP%5C%22%2C%5C%22innersid%5C%22%3A%5C%22%5C%22%2C%5C%22innerouid%5C%22%3A%5C%22%5C%22%2C%5C%22pushcode%5C%22%3A%5C%22%5C%22%2C%5C%22txCpsId%5C%22%3A%5C%22%5C%22%2C%5C%22amsPid%5C%22%3A%5C%22%5C%22%2C%5C%22gdt_vid%5C%22%3A%5C%22%5C%22%7D%22%2C%22scene%22%3A1007%2C%22personalRecommendSwitch%22%3Atrue%2C%22localRecommendSwitch%22%3Atrue%2C%22marketSwitch%22%3Atrue%2C%22pLen%22%3A3%7D&_obt=1697011436515", "Accept-Encoding": "gzip, deflate", "Accept-Language": "zh-CN,zh;q=0.9"}
    burp0_json={"commentTagV2List": [], "commonStatisticList": [1], "ftype": "v", "head": {"aid": "262684", "auth": "", "cid": "52271115296426477628", "ctok": "", "currency": "CNY", "cver": "1695277203395", "device": "WAP", "extension": [{"name": "supportWebP", "value": "true"}, {"name": "channelcode", "value": "wechat-hybrid"}], "group": "CTRIP", "lang": "01", "locale": "zh-CN", "ouid": "", "sid": "711465", "syscode": "09", "vid": "1697008100680.1e1cmmiBfNEU", "xsid": ""}, "hotelId": int(hotelId), "pageIndex": int(pageIndex), "pageSize": 20, "repeatComment": 1, "session": {"fp": "U5zWTbIPwq3i31R4de14v6Orz1wOBYs0EmY99RgGIc0ip5i9SWOyDbeQ3W8YXTyZzKzaeBkezlEOZjctWdki1fYgtynYF9eAnYPTWbmjf4w5ZvDpj9J4ow6TRHMWABjmJs5jNhwTZvZLjNJz5vcnv3fEPcwd7jlNem0itPYhNEq9WU0IsY3SK9NYqtxXtvfnYSGy6Tjc5vUFed5YbZjnfyqFwz6EmGW8YQPI4PeQde1NKMljOkJLBJGtwpUeSTxUXj9zwFtr50Yqoe3YQZEnowXmRQ7Y0LxOLYTYTnJdcWfnrXJzbwhkImYsDjfMKqDjGhv35Yknycnjo8vSmemhYfMj8lyXJfSvA5YTZyoZjq8vpgeQ6YZmjB9ymJ78YoFvmOWGNWTdrMnwn3wSYLmJ1oiPaKq9v9AeShYNniXGYmjbsxa5E1Y8kIq9KQHe4mIdGRXDRUYH8JtAWPXvloEddxS0Y0YgFYhPJPtrq5YMaiPbitfikbj7GYN7jTar3Y83KkaI3tYXUJ0UvSsYtkwL9jtPySLY5zw9FrHpwbZy1YqTr0fImEadRUQwZOvoSyzQW8LRQ5RtTEMqiOlyh9jGQJNfiq3xkZjlcvBYd9ihAvazYgQymcEhFKbYdLrPzITzv9mRfAw0fvHay9hW4XWS3jH1RaqWnMW0Uwdkj15WO7e0OWSlwkYsBwqJ8EpdjSawkgvOtjSnxthEZhjSYTQyDfih6xn4R8FJcdiHawf9eb9vf3wOFWBqw6PWPoeBY5FrgXvAGxofR7gJsBiZUw4aeBpvXfRN9YXMyXvqNwfYUQEmvgqjUBRoBwGZv3OyT5W6HRGPRt1EzbiSPy6Lj9pJ67R8TE6tI40eQYLmi0nEhSJD7EZkjl8WfBW3DW1GYLHY97YnmRNcYHfWoOYUDYadYUBjbHezpEDLW78eoSwTzecOjHMYADyGUEZ7jb8EhqrZtj7HwGDyodwOUWX0EsYSqR9UWn4WOFWXcWntYtYPwcE8hiHTvDoEOaWfFymNjoJOmvFZEdMWPmyGBj9MWp9JLGeUYmFygmjmpR3SElAELDEd7R0nEoGwa3YAoKMYDnKNPKfAIf3EnhEBlE6PY1NYb3YXZYd5ySGKUc", "key": "640dc9260f99aea44464a8Q0753c6a79fb692ej74deddKp9af12oFf72a9e3\xc2\x831`"}}
    res = requests.post(burp0_url, headers=burp0_headers, cookies=burp0_cookies, json=burp0_json)
    data = res.json()

    # 获取"groupList"数组
    group_list = data.get('groupList', [])
    # 遍历"groupList"数组，找到dataType为2的"commentList"数组
    comment_list = []
    for group in group_list:
        if group.get('dataType') == 2:
            comment_list = group.get('commentList', [])
            break
    # 获取comment_list数组中所有元素的content值
    comments = [comment.get('content') for comment in comment_list]
    return comments
"""
with open('ip1.txt', 'r') as file:
    proxy_list = [line.strip() for line in file]
user_agents = [
    "'Mozilla/5.0 (iPhone; CPU iPhone OS 13_2_3 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/13.0.3 Mobile/15E148 Safari/604.1'",
    "'Mozilla/5.0 (Linux; Android 8.0.0; SM-G955U Build/R16NW) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.141 Mobile Safari/537.36'",
    "'Mozilla/5.0 (Linux; Android 10; SM-G981B) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.162 Mobile Safari/537.36'",
    "'Mozilla/5.0 (iPad; CPU OS 13_3 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) CriOS/87.0.4280.77 Mobile/15E148 Safari/604.1'",
    "'Mozilla/5.0 (Linux; Android 8.0; Pixel 2 Build/OPD3.170816.012) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/102.0.0.0 Mobile Safari/537.36'",
    "'Mozilla/5.0 (Linux; Android) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/88.0.4324.109 Safari/537.36 CrKey/1.54.248666'",
    "'Mozilla/5.0 (X11; Linux aarch64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/88.0.4324.188 Safari/537.36 CrKey/1.54.250320'",
    "'Mozilla/5.0 (BB10; Touch) AppleWebKit/537.10+ (KHTML, like Gecko) Version/10.0.9.2372 Mobile Safari/537.10+'",
    "'Mozilla/5.0 (PlayBook; U; RIM Tablet OS 2.1.0; en-US) AppleWebKit/536.2+ (KHTML like Gecko) Version/7.2.1.0 Safari/536.2+'",
    "'Mozilla/5.0 (Linux; U; Android 4.3; en-us; SM-N900T Build/JSS15J) AppleWebKit/534.30 (KHTML, like Gecko) Version/4.0 Mobile Safari/534.30'",
    "'Mozilla/5.0 (Linux; U; Android 4.1; en-us; GT-N7100 Build/JRO03C) AppleWebKit/534.30 (KHTML, like Gecko) Version/4.0 Mobile Safari/534.30'",
    "'Mozilla/5.0 (Linux; U; Android 4.0; en-us; GT-I9300 Build/IMM76D) AppleWebKit/534.30 (KHTML, like Gecko) Version/4.0 Mobile Safari/534.30'",
    "'Mozilla/5.0 (Linux; Android 7.0; SM-G950U Build/NRD90M) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/62.0.3202.84 Mobile Safari/537.36'",
    "'Mozilla/5.0 (Linux; Android 8.0.0; SM-G965U Build/R16NW) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.111 Mobile Safari/537.36'",
    "'Mozilla/5.0 (Linux; Android 8.1.0; SM-T837A) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/70.0.3538.80 Safari/537.36'",
    "'Mozilla/5.0 (Linux; U; en-us; KFAPWI Build/JDQ39) AppleWebKit/535.19 (KHTML, like Gecko) Silk/3.13 Safari/535.19 Silk-Accelerated=true'",
    "'Mozilla/5.0 (Linux; U; Android 4.4.2; en-us; LGMS323 Build/KOT49I.MS32310c) AppleWebKit/537.36 (KHTML, like Gecko) Version/4.0 Chrome/102.0.0.0 Mobile Safari/537.36'",
    "'Mozilla/5.0 (Windows Phone 10.0; Android 4.2.1; Microsoft; Lumia 550) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/46.0.2486.0 Mobile Safari/537.36 Edge/14.14263'",
    "'Mozilla/5.0 (Linux; Android 6.0.1; Moto G (4)) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/102.0.0.0 Mobile Safari/537.36'",
    "'Mozilla/5.0 (Linux; Android 6.0.1; Nexus 10 Build/MOB31T) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/102.0.0.0 Safari/537.36'",
    "'Mozilla/5.0 (Linux; Android 4.4.2; Nexus 4 Build/KOT49H) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/102.0.0.0 Mobile Safari/537.36'",
    "'Mozilla/5.0 (Linux; Android 6.0; Nexus 5 Build/MRA58N) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/102.0.0.0 Mobile Safari/537.36'",
    "'Mozilla/5.0 (Linux; Android 8.0.0; Nexus 5X Build/OPR4.170623.006) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/102.0.0.0 Mobile Safari/537.36'",
    "'Mozilla/5.0 (Linux; Android 7.1.1; Nexus 6 Build/N6F26U) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/102.0.0.0 Mobile Safari/537.36'",
    "'Mozilla/5.0 (Linux; Android 8.0.0; Nexus 6P Build/OPP3.170518.006) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/102.0.0.0 Mobile Safari/537.36'",
    "'Mozilla/5.0 (Linux; Android 6.0.1; Nexus 7 Build/MOB30X) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/102.0.0.0 Safari/537.36'",
    "'Mozilla/5.0 (compatible; MSIE 10.0; Windows Phone 8.0; Trident/6.0; IEMobile/10.0; ARM; Touch; NOKIA; Lumia 520)'",
    "'Mozilla/5.0 (MeeGo; NokiaN9) AppleWebKit/534.13 (KHTML, like Gecko) NokiaBrowser/8.5.0 Mobile Safari/534.13'",
    "'Mozilla/5.0 (Linux; Android 9; Pixel 3 Build/PQ1A.181105.017.A1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/66.0.3359.158 Mobile Safari/537.36'",
    "'Mozilla/5.0 (Linux; Android 10; Pixel 4) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/81.0.4044.138 Mobile Safari/537.36'",
    "'Mozilla/5.0 (Linux; Android 11; Pixel 3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/88.0.4324.181 Mobile Safari/537.36'",
    "'Mozilla/5.0 (Linux; Android 5.0; SM-G900P Build/LRX21T) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/102.0.0.0 Mobile Safari/537.36'",
    "'Mozilla/5.0 (Linux; Android 8.0; Pixel 2 Build/OPD3.170816.012) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/102.0.0.0 Mobile Safari/537.36'",
    "'Mozilla/5.0 (Linux; Android 8.0.0; Pixel 2 XL Build/OPD1.170816.004) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/102.0.0.0 Mobile Safari/537.36'",
    "'Mozilla/5.0 (iPhone; CPU iPhone OS 10_3_1 like Mac OS X) AppleWebKit/603.1.30 (KHTML, like Gecko) Version/10.0 Mobile/14E304 Safari/602.1'",
    "'Mozilla/5.0 (iPhone; CPU iPhone OS 13_2_3 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/13.0.3 Mobile/15E148 Safari/604.1'",
    "'Mozilla/5.0 (iPad; CPU OS 11_0 like Mac OS X) AppleWebKit/604.1.34 (KHTML, like Gecko) Version/11.0 Mobile/15A5341f Safari/604.1'",
]

def select_random_proxy(proxy_list):
    return random.choice(proxy_list)


"""
def get_comments(hotelId,pageIndex):
    global user_agents
    random_proxy = select_random_proxy(proxy_list)
    ip, port, username, password = random_proxy.split(':')
    proxy_url = f"http://{username}:{password}@{ip}:{port}"
    proxies = {
        'http': proxy_url,
        'https': proxy_url
    }
    import requests

    burp0_url = "https://m.ctrip.com:443/restapi/soa2/24626/commentlist?_fxpcqlniredt=52271165296426527463&x-traceID=52271165296426527463-1697485763159-7436135"
    burp0_cookies = {"GUID": "52271165296426527463", "nfes_isSupportWebP": "1", "nfes_isSupportWebP": "1",
                     "UBT_VID": "1697185638111.5289MueHV1yF", "_RF1": "130.88.226.10", "_RSG": "7dKmf19p5FClVApH2K_Fx8",
                     "_RDG": "28ac8579459d0423e21eb31df7882b69c8", "_RGUID": "c3c5cc90-69c1-49c7-8ecf-fd4091cca7ef",
                     "MKT_Pagesource": "H5",
                     "_ubtstatus": "%7B%22vid%22%3A%221697185638111.5289MueHV1yF%22%2C%22sid%22%3A2%2C%22pvid%22%3A4%2C%22pid%22%3A0%7D",
                     "_bfaStatusPVSend": "1", "_bfi": "p1%3D228029%26p2%3D10650056194%26v1%3D4%26v2%3D3",
                     "_bfaStatus": "success", "librauuid": "", "hotelhst": "1164390341",
                     "cticket": "E3C3F84B6EAA3B312FBF741D5D7E7C1DF90A8B90A4512FE3CFEA91CF4D54F2F1", "login_type": "0",
                     "login_uid": "1FE3F2E9E342DAAF827E00B07E412196", "DUID": "u=2A73E2E2758F5C57A5EED83C25C6EC55&v=0",
                     "IsNonUser": "F",
                     "AHeadUserInfo": "VipGrade=0&VipGradeName=%C6%D5%CD%A8%BB%E1%D4%B1&UserName=&NoReadMessageCount=0",
                     "_pd": "%7B%22_o%22%3A3%2C%22s%22%3A183%2C%22_s%22%3A0%7D",
                     "_bfa": "1.1697185638111.5289MueHV1yF.1.1697485761510.1697485762263.3.5.228032",
                     "Union": "OUID=mini1053&AllianceID=1314167&SID=4258862&SourceID=55555549&createtime=1697485762&Expires=1698090562329",
                     "MKT_OrderClick": "ASID=13141674258862&AID=1314167&CSID=4258862&OUID=mini1053&CT=1697485762330&CURL=https%3A%2F%2Fm.ctrip.com%2Fwebapp%2Fhotels%2Fcommentlist%3Fid%3D49212463%26hotelname%3D%25E5%25B7%25B4%25E5%25A7%2586%25E5%2593%25A5%25E5%25BA%25A6%25E5%2581%2587%25E6%259D%2591(Pamookkoo)%26biz%3D2%26ftype%3Dv%26fromminiapp%3Dweixin%26allianceid%3D1314167%26sid%3D4258862%26ouid%3Dmini1053%26sourceid%3D55555549%26_cwxobj%3D%257B%2522cid%2522%253A%252252271165296426527463%2522%252C%2522appid%2522%253A%2522wx0e6ed4f51db9d078%2522%252C%2522mpopenid%2522%253A%252233d6ec3b-d887-4cf8-8fc4-c29c77f47342%2522%252C%2522mpunionid%2522%253A%2522oHkqHt7GNBnAyCarN1KtQAnQhKPI%2522%252C%2522allianceid%2522%253A%25221314167%2522%252C%2522sid%2522%253A%25224258862%2522%252C%2522ouid%2522%253A%2522mini1053%2522%252C%2522sourceid%2522%253A%252255555549%2522%252C%2522exmktID%2522%253A%2522%257B%255C%2522openid%255C%2522%253A%255C%252233d6ec3b-d887-4cf8-8fc4-c29c77f47342%255C%2522%252C%255C%2522unionid%255C%2522%253A%255C%2522oHkqHt7GNBnAyCarN1KtQAnQhKPI%255C%2522%252C%255C%2522channelUpdateTime%255C%2522%253A%255C%25221697485177892%255C%2522%252C%255C%2522serverFrom%255C%2522%253A%255C%2522WAP%252FWECHATAPP%255C%2522%252C%255C%2522innersid%255C%2522%253A%255C%2522%255C%2522%252C%255C%2522innerouid%255C%2522%253A%255C%2522%255C%2522%252C%255C%2522pushcode%255C%2522%253A%255C%2522%255C%2522%252C%255C%2522txCpsId%255C%2522%253A%255C%2522%255C%2522%252C%255C%2522amsPid%255C%2522%253A%255C%2522%255C%2522%252C%255C%2522gdt_vid%255C%2522%253A%255C%2522%255C%2522%257D%2522%252C%2522scene%2522%253A1053%252C%2522personalRecommendSwitch%2522%253Atrue%252C%2522localRecommendSwitch%2522%253Atrue%252C%2522marketSwitch%2522%253Atrue%252C%2522pLen%2522%253A5%257D%26_obt%3D1697485680850&VAL={\"pc_vid\":\"1697185638111.5289MueHV1yF\"}"}
    burp0_headers = {"Sec-Ch-Ua": "\"Chromium\";v=\"107\", \"Not=A?Brand\";v=\"24\"",
                     "Cookieorigin": "https://m.ctrip.com", "Content-Type": "application/json",
                     "Sec-Ch-Ua-Mobile": "?0",
                     "User-Agent": random.choice(user_agents),
                     "Sec-Ch-Ua-Platform": "\"Windows\"", "Accept": "*/*", "Origin": "https://m.ctrip.com",
                     "Sec-Fetch-Site": "same-origin", "Sec-Fetch-Mode": "cors", "Sec-Fetch-Dest": "empty",
                     "Referer": "https://m.ctrip.com/webapp/hotels/commentlist?id=49212463&hotelname=%E5%B7%B4%E5%A7%86%E5%93%A5%E5%BA%A6%E5%81%87%E6%9D%91(Pamookkoo)&biz=2&ftype=v&fromminiapp=weixin&allianceid=1314167&sid=4258862&ouid=mini1053&sourceid=55555549&_cwxobj=%7B%22cid%22%3A%2252271165296426527463%22%2C%22appid%22%3A%22wx0e6ed4f51db9d078%22%2C%22mpopenid%22%3A%2233d6ec3b-d887-4cf8-8fc4-c29c77f47342%22%2C%22mpunionid%22%3A%22oHkqHt7GNBnAyCarN1KtQAnQhKPI%22%2C%22allianceid%22%3A%221314167%22%2C%22sid%22%3A%224258862%22%2C%22ouid%22%3A%22mini1053%22%2C%22sourceid%22%3A%2255555549%22%2C%22exmktID%22%3A%22%7B%5C%22openid%5C%22%3A%5C%2233d6ec3b-d887-4cf8-8fc4-c29c77f47342%5C%22%2C%5C%22unionid%5C%22%3A%5C%22oHkqHt7GNBnAyCarN1KtQAnQhKPI%5C%22%2C%5C%22channelUpdateTime%5C%22%3A%5C%221697485177892%5C%22%2C%5C%22serverFrom%5C%22%3A%5C%22WAP%2FWECHATAPP%5C%22%2C%5C%22innersid%5C%22%3A%5C%22%5C%22%2C%5C%22innerouid%5C%22%3A%5C%22%5C%22%2C%5C%22pushcode%5C%22%3A%5C%22%5C%22%2C%5C%22txCpsId%5C%22%3A%5C%22%5C%22%2C%5C%22amsPid%5C%22%3A%5C%22%5C%22%2C%5C%22gdt_vid%5C%22%3A%5C%22%5C%22%7D%22%2C%22scene%22%3A1053%2C%22personalRecommendSwitch%22%3Atrue%2C%22localRecommendSwitch%22%3Atrue%2C%22marketSwitch%22%3Atrue%2C%22pLen%22%3A5%7D&_obt=1697485680850",
                     "Accept-Encoding": "gzip, deflate", "Accept-Language": "zh-CN,zh;q=0.9"}
    burp0_json = {"commentTagV2List": [], "commonStatisticList": [1], "ftype": "v",
                  "head": {"aid": "1314167", "auth": "", "cid": "52271165296426527463", "ctok": "", "currency": "CNY",
                           "cver": "1697114451679", "device": "WAP",
                           "extension": [{"name": "supportWebP", "value": "true"},
                                         {"name": "channelcode", "value": "wechat-hybrid"}], "group": "CTRIP",
                           "lang": "01", "locale": "zh-CN", "ouid": "mini1053", "sid": "4258862", "syscode": "09",
                           "vid": "1697185638111.5289MueHV1yF", "xsid": ""}, "hotelId": int(hotelId), "pageIndex": int(pageIndex),
                  "pageSize": 10, "repeatComment": 1, "session": {
            "fp": "p0nIl1Rp7IHlrDcrhoenMEtTWnAw54Y3OeGYAXiPFwB3YMHYd7EgGJ7pvaSvAYckeQkYqmiztekdEmljQ5WpQesSWdZrHYDPwlHyUtI8Oysbw01r1oKGJ9ARGBrzGv8pr3JmZjT9wpOvqAjlJPavUHvFcYzNwTSjsGeg5inMYp9jz6xtPY4YNoxUyksyQTv5HY6hy04j5zv3deXUYl5jzNy7SRGOWDSKhYfqYL5rNsvZnKs8jnQJBsJdPwkne9Sx4pjDMwGMWkcJSli4YXaWz7iQvLNrgaY9fj6YH3W8cwXAeZtw35yq8ynYsUwZmekzY1avn5YmhyDNjhTvaSedfYk7jaAyTJcXvmnYN0yg4jqPv7feLqYH5jZQymJ35Y4ovgXWOtWL3RUdEmFxtYH6jljmHYZbvqleL9YXqiQsYsMWPMWM1eDYNbWsnE5PWq7KhFWB9WlYkfyUv4dYSLvDDxPQKdYUUKBFIz4WAGYsAigZilfi4QjBgEASxzwsYAMEmhv6tEnzJfnv6zYXzwZtjpQy3gYBQwdJGDjOSY1YqvmSrLGwnzR3Uw8Av7zydZEatj79Jfmy03whbR4gJ9TJo7wBpeqGemPwkYLUEphYMbK3Oi5ajHwDYzDWcaY4GrObRmNwsdvPLyQTR9fj3MJ5LwPqiUpjMcR1mRL8RzawF8KXMYoYs4x86wlPe1Xj6qwLUvSgj3cKHgKFoxQYA7xnzYlNy34RNnJDci4cwF0e5pjthwL4Ed9IcHwmyHYGqyzpRbQr3ORoTJGni6awaUegnjlqRhAw7gEaDY9ZIgYSTwzni0hetmRmswgcvf9yD0ETtjk4J8MyL1w1mRt8Jh4JGHJmmKlyz5y6YnOE8twtOrf4EQmjlSWsNWaBWmHY8UYsoY0qRGbYhzWodYfQYc1Y0MjsUeTPEp4Whce3pw1oeTsjhFYpOyPdEa5j1mEd1rLqjzqw9ky4BY1OKP4KtYGSRFpWTbWUTW9oWlTYSY65ED3iHQKXMvk7ElPWmLyHmj8JfbvGmEp4WSpyq9jqfEnoEDSvLYZ4Yc7R9qvs7EqDEBQEH7R4HEbpJkqy37wzYm3rnqvNJSmED1EdgEPcYkTYlnY5cY4AimmxNO",
            "key": "b67883bab13dde3\xc2\x80f85f76Q09e5521994b6b05b8d1d3a33368f0c414f22b1f7a"}}
    time.sleep(5)
    res = requests.post(burp0_url, headers=burp0_headers, cookies=burp0_cookies, json=burp0_json,proxies=proxies)
    print("这次用了"+burp0_headers["User-Agent"]+"请求头")
    print(f"评论信息使用{proxies}爬取")
    time.sleep(3)
    data = res.json()
    grouplist = data.get("groupList", [])
    comment_data = {'comments_with_images': []}
    for group in grouplist:
        if group.get('dataType') == 2:
            comment_list = group.get('commentList', [])

            # 使用一个空字符串初始化 comment_content
            comment_content = ""

            # 遍历评论列表中的每个评论
            for comment in comment_list:
                # 获取评论内容
                comment_content = comment.get('content', comment_content)

                # 获取评论中的图片信息
                image_urls = []
                if 'imageCuttingsList' in comment:
                    image_cuttings = comment['imageCuttingsList']
                    image_urls = [image_cutting.get('bigImageUrl', "") for image_cutting in image_cuttings]

                # 创建包含评论和图片链接的字典
                comment_with_images = {
                    'comment_content': comment_content,
                    'image_urls': image_urls
                }

                # 将评论和图片链接的字典添加到列表中
                comment_data['comments_with_images'].append(comment_with_images)

    return comment_data
"""

"""
    for group in grouplist:
        if group.get('dataType') == 2:
            comment_list = group.get('commentList', [])

            # 遍历评论列表中的每个评论
            for comment in comment_list:
                # 获取评论内容
                comment_content = comment.get('content')

                # 获取评论中的图片信息
                if 'imageCuttingsList' in comment:
                    image_cuttings = comment['imageCuttingsList']
                    image_urls = [image_cutting.get('bigImageUrl') for image_cutting in image_cuttings]

                    # 创建包含评论和图片链接的字典
                    comment_with_images = {
                        'comment_content': comment_content,
                        'image_urls': image_urls
                    }

                    # 将评论和图片链接的字典添加到列表中
                    comment_data['comments_with_images'].append(comment_with_images)
    return comment_data
"""
def get_comments(hotelId,pageindex):
    global user_agents
    burp0_url = "https://uk.trip.com:443/restapi/soa2/28820/getCommentList?testab=2231a88c3402e2e5ee83d32fd80964067141691960db0a5a819bcad5fe92f084&x-traceID=1698077771987.63a4QEc5sdbB-1698080722015-1039732543"
    burp0_cookies = {
        "Union": "AllianceID=1052761&SID=1816122&OUID=ctag.hash.dd8957658385&Expires=1700669771986&createtime=1698077771",
        "UBT_VID": "1698077771987.63a4QEc5sdbB",
        "kafka_result": "{\"isDirectVisit\":\"0\",\"hasUrlLocale\":\"1\",\"hasCookieLocale\":\"0\",\"isUrlCookieSame\":\"0\",\"isJump\":\"0\",\"jumpType\":\"targetLocale\",\"platform\":\"online\"}",
        "ibu_online_home_language_match": "{\"isRedirect\":false,\"isShowSuggestion\":false,\"lastVisited\":true,\"region\":\"gb\",\"redirectSymbol\":false,\"site_url\":[]}",
        "ibulanguage": "EN", "ibulocale": "en_gb", "cookiePricesDisplayed": "GBP",
        "_abtest_userid": "275b5bfb-dc22-41fe-8cde-383654a6d102", "ibu_oh_last_visited_at": "1698077772893",
        "_gid": "GA1.2.741657148.1698077773",
        "_gac_UA-109672825-1": "1.1698077773.CjwKCAjws9ipBhB1EiwAccEi1A7EeiaGhxfzL6QFbO229cRl4f--fVP_IE4f3NFKOK8nx7ICOmWjIRoCpWcQAvD_BwE",
        "_gcl_aw": "GCL.1698077774.CjwKCAjws9ipBhB1EiwAccEi1A7EeiaGhxfzL6QFbO229cRl4f--fVP_IE4f3NFKOK8nx7ICOmWjIRoCpWcQAvD_BwE",
        "_gcl_dc": "GCL.1698077774.CjwKCAjws9ipBhB1EiwAccEi1A7EeiaGhxfzL6QFbO229cRl4f--fVP_IE4f3NFKOK8nx7ICOmWjIRoCpWcQAvD_BwE",
        "_gcl_au": "1.1.1918006855.1698077774",
        "_gac_UA-109672825-3": "1.1698077774.CjwKCAjws9ipBhB1EiwAccEi1A7EeiaGhxfzL6QFbO229cRl4f--fVP_IE4f3NFKOK8nx7ICOmWjIRoCpWcQAvD_BwE",
        "_gac_UA-109672825-13": "1.1698077774.CjwKCAjws9ipBhB1EiwAccEi1A7EeiaGhxfzL6QFbO229cRl4f--fVP_IE4f3NFKOK8nx7ICOmWjIRoCpWcQAvD_BwE",
        "_combined": "transactionId%3Dba99823edb262b2de08e4e432fb75219", "_tt_enable_cookie": "1",
        "_ttp": "w31iVsDoEPj22rkzwlkQ2ps8HPO", "_tp_search_latest_channel_name": "hotels", "_RF1": "130.88.226.10",
        "_RSG": "eVaClGMyB2DAnkG2IX81z8", "_RDG": "28bfcb44b1dbf72dcd1f3dc2a39ba5226d",
        "_RGUID": "d08bf164-2f32-47e8-8f41-22148242ec4a", "g_state": "{\"i_p\":1698084986976,\"i_l\":1}",
        "IBU_TRANCE_LOG_P": "56702092987",
        "IBU_TRANCE_LOG_URL": "%2Fhotels%2Flist%3Fcity%3D359%26cityName%3D%25E6%259B%25BC%25E8%25B0%25B7%26provinceId%3D0%26countryId%3D4%26districtId%3D0%26checkin%3D2023%2F10%2F24%26checkout%3D2023%2F11%2F21%26barCurr%3DGBP%26searchType%3DCT%26searchWord%3D%25E6%259B%25BC%25E8%25B0%25B7%26searchValue%3D19%257C359_19_359_1%26searchCoordinate%3DBAIDU_-1_-1_0%257CGAODE_-1_-1_0%257CGOOGLE_-1_-1_0%257CNORMAL_13.7278956_100.5241235_0%26crn%3D1%26adult%3D2%26children%3D0%26searchBoxArg%3Dt%26travelPurpose%3D0%26ctm_ref%3Dix_sb_dl%26domestic%3Dfalse",
        "devicePixelRatio": "1.25", "oldLocale": "en-GB", "_fbp": "fb.1.1698077794614.1403447348",
        "_bfaStatusPVSend": "1", "ibu_online_permission_cls_ct": "1", "ibu_online_permission_cls_gap": "1698077811988",
        "htltmp": "", "htlstmp": "", "librauuid": "",
        "ibu_hotel_search_date": "%7B%22checkIn%22%3A%222023%2F10%2F24%22%2C%22checkOut%22%3A%222023%2F11%2F21%22%7D",
        "intl_ht1": "h4%3D725_998454%2C359_1457601", "hotel": "998454", "IBU_showtotalamt": "3",
        "hotelhst": "1164390341", "_gat_UA-109672825-3": "1", "_gat": "1",
        "_ga_2DCSB93KS4": "GS1.2.1698077774.1.1.1698080677.60.0.0",
        "_ga_37RNVFDP1J": "GS1.2.1698077773.1.1.1698080677.60.0.0",
        "_bfi": "p1%3D10320668147%26p2%3D10320668147%26v1%3D14%26v2%3D10",
        "_bfa": "1.1698077771987.63a4QEc5sdbB.1.1698080675276.1698080720654.1.15.10320668147",
        "_ubtstatus": "%7B%22vid%22%3A%221698077771987.63a4QEc5sdbB%22%2C%22sid%22%3A1%2C%22pvid%22%3A15%2C%22pid%22%3A10320668147%7D",
        "_bfaStatus": "send", "_uetsid": "88e6c99071bf11ee80c0899bb686cf34",
        "_uetvid": "88e6ed6071bf11eea429b3e2d9511144", "wcs_bt": "s_33fb334966e9:1698080721",
        "_ga_X437DZ73MR": "GS1.1.1698077774.1.1.1698080721.0.0.0", "_ga": "GA1.2.1874431117.1698077773"}
    burp0_headers = {"Sec-Ch-Ua": "\"Chromium\";v=\"118\", \"Google Chrome\";v=\"118\", \"Not=A?Brand\";v=\"99\"",
                     "Currency": "GBP", "Locale": "en-GB", "Sec-Ch-Ua-Mobile": "?0",
                     "User-Agent": random.choice(user_agents),
                     "Trip-Trace-Id": "1698077771987.63a4QEc5sdbB-1698080722015-1039732543",
                     "Content-Type": "application/json", "Accept": "application/json",
                     "Pid": "1f8ed06b-1918-46e0-8fed-db7f1d2f1664", "P": "56702092987",
                     "X-Traceid": "1698077771987.63a4QEc5sdbB-1698080722015-1039732543",
                     "Sec-Ch-Ua-Platform": "\"Windows\"", "Origin": "https://uk.trip.com",
                     "Sec-Fetch-Site": "same-origin", "Sec-Fetch-Mode": "cors", "Sec-Fetch-Dest": "empty",
                     "Referer": "https://uk.trip.com/hotels/detail/?cityId=725&hotelId=998454&checkIn=2023-10-24&checkOut=2023-11-21&adult=2&children=0&subStamp=1638&crn=1&ages=&travelpurpose=0&curr=GBP&link=title&hoteluniquekey=H4sIAAAAAAAAAOM6LcjFJMEkxMTBKHVAkKP_yekGVou9Uo5vAnfItar2OngaA8Fm13sOATyTGN05H1aJrDtesN5BkAEETqx0UNLh2L9mgYSAlsT9w812CqyaYAkGLQdDGMOiVTCIlWMiowRLFIMTK8c1NwmWSYyelJqkwUhFs5ioaBYzFc1ioaJZrFQ0i42KZrHDzfLgPPxVI6b_4naoWR2bHZS0sZplAjPLBGwWC4eYBBPQKA6gUd3ZVDGKk3pGcVHPKG7qGcVDPaN4qWcUH_WM4qeeUQLUM0qQekYJUc8oYeoZJUI9o0SpZ5QY9YwSp55REtQzSpJ6RklRzyhpqFEzGLf9sNnIOGsmEAhvdtjByHyAMeQEY5f0AqbOtVOZdzGxcXyZLCrBcgjIeLKKWYLlFBMrxxoWCaZLTAy3mBgeMTG8YmL4xMTwCyh8O0GCpYmZleMYsProYmaYBGTO55dgmcXMsIiZQYo3LdkwxcAiOTEtycwySUFI4-GBWyvZjJQmMTK5O51ilDI0s7QwMAcCQ0sLcz0z40STQNdk0-KUJCcrZilGNw-ZIDY3ZzNXS6coLS7mEA8niMcZPthLMbs7BSjara94NZFtur0WSM6w7MEcwaV7T9onsabm6bo7ZdzlLWBsYGTuYuQQYPRgjGCsYHzFCNL2A-R9ADjAsAFpCQAA&subChannel=&masterhotelid_tracelogid=fc1d08cafb69b&NewTaxDescForAmountshowtype0=F&barcurr=GBP",
                     "Accept-Encoding": "gzip, deflate", "Accept-Language": "zh-CN,zh;q=0.9"}
    burp0_json = {"commentTagList": [], "commentTagV2List": [], "commonStatisticList": [],
                  "functionOptions": ["IntegratedTARating", "hidePicAndVideoAgg", "TripReviewsToServerOnline",
                                      "IntegratedExpediaList"],
                  "head": {"aid": "1052761", "bu": "Hotel", "cid": "1698077771987.63a4QEc5sdbB", "currency": "GBP",
                           "cver": "0", "group": "trip", "guid": "", "isSSR": False, "locale": "en-GB",
                           "ouid": "ctag.hash.dd8957658385", "pageId": "10320668147", "platform": "PC",
                           "sid": "1816122", "timezone": "8", "vid": "1698077771987.63a4QEc5sdbB"}, "hotelExtension": {
            "fingerprint": "TqDjOXYB3vHdrdNeB7eq4E3BY6PjsPEBQvlYtyPLwzkwbgYsoEGcYgci3XJlYg4KzBK4dWBZeXgEgFjqGWA7Woai7bwHYc4IhvDdJkqjd4wkdv5FyoJ7diFcWacv08ytJQUjaAwnPvS9jhJ8cvLMvcoYX6wdUjXQeoSiT8Y1JsAvSqiUYmAe8NEQ0jhAyUpj7GvbLE8ZvUhWO0jzSiqoyzbxhYMtwg8YTcY3MYTBi9Nwm9RHAE9zWgMeFpIB8vUYHSr1ge08vAPJ1Owm3JdYNGx93yhHxlXEhTxq9yzYXHIGdWLyozvp9YbLydljO7vULeb6YBkj3FyhJNAvBLYqoyHdjdPv57ehgYQ9jhBylJUZYm0v56WM5W9TyfPYHHKdYcBJhnYdcidPvDse5cY5HiG8YHmEQUiM5RbYanYUPeB9vQkjBBxmseOYALyoPjp7ET0j7kIsvXYh9y3tIq6woHYLqilfi8Mi0Aj0bxLME0QjBYzDwt7JfhJNqylfwfoY6swSgyMTwb9YUmwtSw09JSnesYD1js6ekqj5HRq4wLDv7mjnBWXhj5kWTLy5SYp3WmoJ63Jh7i4TI6SI3cYFY40icXrOymti58JZ9ylYG4RNbIaAYUDRs0wzcv7ojD7WmTyo3yFoy0nyfZR54vbDjdBya1RQzyBNJFYMj6LRB1Wo1j5nwaXvlqjaUKAOjZ7rMYD6yfORnhIGLRckJDgifqwH6ehTj8Lw3NENoxTFE8TrfYZlvmfW3hWP8RgMJTDif9w45eTXjGfRcmwa4ipdYfXWoYtXEp7iZfipmR5kwZDvB1j7zWkOjQzWbZyzXYPpWfpJcQJdTYAsWMgeLBJcYz6jtbx6dEhBEHpjdBWBqW36WBPY85Y7HYZSRPzYGUWHgYlSYzpYoDjfTe8cE4pWD4en7wocehZjnkYg4ymsE8bjSaE1Pr5TjXTw3oy0QiUXEzle1YfFRNPWDTWfZWS4W6NYzYgAxmE8pJcQvmGEbGW53yUGjpJPkvbXEAFWZ7ymNj70wBoIGXxNYlSYmFI7Tvp7E7HEcAE4FRPaEdj7owXaRaY4gEDNeXcekdEmQEPBE5tYMBYD4YFqYALxc6YbL",
            "token": "2231a88c3402e2e5ee83d32fd80964067141691960db0a5a819bcad5fe92f084"}, "hotelId": int(hotelId),
                  "languageTypeList": [], "orderBy": 0, "packageList": [], "pageIndex": int(pageindex), "pageSize": 10,
                  "repeatComment": 1, "roomList": [], "travelTypeList": [], "UnusefulReviewPageIndex": 1}
    res = requests.post(burp0_url, headers=burp0_headers, cookies=burp0_cookies, json=burp0_json)
    data = res.json()
    print("这是第" + str(pageindex) + "页的评论")
    # print(data)
    # 提取评论列表
    comment_list = data['data']['commentList']
    # "imageDomain":"ak-d.tripcdn.com"
    imageDomain = data['data']['imageDomain']
    print(imageDomain)
    # 创建一个新字典来存储提取的数据
    result_dict = {}

    # 遍历评论列表并提取内容和图片列表
    for comment in comment_list:
        content = comment['content']
        image_list = comment['imageList']

        # 在每个图片链接前面添加imageDomain和"/image/"前缀
        modified_image_list = [f"{imageDomain}/images{image}" for image in image_list]

        # 将内容和修改后的图片列表存储在新字典中
        result_dict[comment['id']] = {
            'content': content,
            'imageList': modified_image_list
        }
    return result_dict

def get200comments(hotelId):
    comment_list = []
    image_urls_list = []
    total_pages = 30
    current_page = 1
    comment_content = ""  # 初始化 comment_content
    while len(comment_list) < 200 and current_page <= total_pages:
        comment_data = get_comments(hotelId, current_page)
        for item in comment_data.values():
            comment_content = item.get('content', '')  # 使用item.get来获取内容，如果不存在则使用空字符串
            comment_list.append(comment_content)
            image_urls = item.get('imageList', [])  # 使用item.get来获取图片列表，如果不存在则使用空列表
            image_urls_list.append(image_urls)
        current_page += 1
    return comment_list, comment_content, image_urls_list

def get_addressApic(hotelId):
    global user_agents
    random_proxy = select_random_proxy(proxy_list)
    ip, port, username, password = random_proxy.split(':')
    proxy_url = f"http://{username}:{password}@{ip}:{port}"
    proxies = {
        'http': proxy_url,
        'https': proxy_url
    }
    burp0_url = "https://m.ctrip.com:443/restapi/soa2/26187/graphql?_fxpcqlniredt=52271115296426477628"
    burp0_cookies = {"Union": "OUID=&AllianceID=262684&SID=711465&SourceID=55552689", "DUID": "u=B34F49EE0A13E8463539186D23B00A8B&v=0", "GUID": "52271115296426477628"}
    burp0_headers = {"X-Ctx-Locale": "zh-CN", "User-Agent": random.choice(user_agents), "X-Ctx-Group": "ctrip", "Content-Type": "application/json", "X-Wx-Openid": "b9787617-697b-4fe6-9942-f5fddd4b8063", "X-Ctx-Personal-Recommend": "1", "Xweb_xhr": "1", "X-Ctx-Region": "CN", "X-Ctx-Currency": "CNY", "Duid": "u=B34F49EE0A13E8463539186D23B00A8B&v=0", "Accept": "*/*", "Sec-Fetch-Site": "cross-site", "Sec-Fetch-Mode": "cors", "Sec-Fetch-Dest": "empty", "Referer": "https://servicewechat.com/wx0e6ed4f51db9d078/800/page-frame.html", "Accept-Encoding": "gzip, deflate", "Accept-Language": "zh-CN,zh;q=0.9"}
    burp0_json={"head": {"auth": "8A38C2D39415D5002658D2670FDD4F8D369AEE5E25A10B5F129505F540395603", "cid": "52271115296426477628", "ctok": "", "cver": "1.1.188", "extension": [{"name": "sdkversion", "value": "3.0.0"}, {"name": "openid", "value": "b9787617-697b-4fe6-9942-f5fddd4b8063"}, {"name": "pageid", "value": "10320654891"}, {"name": "supportWebP", "value": "true"}, {"name": "ubt", "value": "{\"vid\":\"1697008079504.6vx93z\",\"sid\":1,\"pvid\":38,\"ts\":1697011325377,\"create\":1697008079504,\"pid\":\"10320654891\"}"}, {"name": "supportFuzzyPrice", "value": "1"}, {"name": "appId", "value": "wx0e6ed4f51db9d078"}, {"name": "scene", "value": "1007"}], "lang": "01", "sauth": "", "sid": "", "syscode": "30"}, "query": " { hotel(id: "+str(hotelId)+", checkIn: \"2023-10-11\", checkOut: \"2023-10-12\") { getBaseInfo { hotelName hotelEnName zoneName address openYear fitmentYear fuzzyAddressTip commentScore commentDesc commentCount bestCommentSentence isOversea cityId cityName totalPictureCount mgrGroupId hotelCategoryOutlineImages { categoryName pictureList { url urlBody urlExtend } } coordinate { latitude longitude } starInfo { star } topAwardInfo { listSubTitle listUrl awardIconUrl lableId rankId annualListAwardIconUrl annualListTagUrl } } getTrafficDetail(filterValue: \"\") { defaultTrafficText } getDetailTag { starTag { icon } dStarTag { icon } medalTag { icon } primeTag { icon } facilityTags(limit: 3) { title } categoryTag { title } } } } ", "source": "hotel_detail_head"}
    time.sleep(5)
    res = requests.post(burp0_url, headers=burp0_headers, cookies=burp0_cookies, json=burp0_json, proxies=proxies)
    print(f"酒店地址图片使用{proxies}爬取")
    print("这次用了"+burp0_headers["User-Agent"]+"请求头")
    time.sleep(5)
    # 将响应内容解析为JSON格式
    data = res.json()
    # print(data)
    # 提取数据
    # Extract data
    address = data['data']['hotel']['getBaseInfo'].get('address', "")
    picList = data['data']['hotel']['getBaseInfo'].get('hotelCategoryOutlineImages', [])
    picList = [item.get('pictureList', []) for item in picList if 'pictureList' in item]
    pictureUrls = [item.get('url', "") for sublist in picList for item in sublist if 'url' in item]
    urlBodies = [item.get('urlBody', "") for sublist in picList for item in sublist if 'urlBody' in item]
    urlExtends = [item.get('urlExtend', "") for sublist in picList for item in sublist if 'urlExtend' in item]
    url_prefix = pictureUrls[0].rsplit('/', 1)[0] + '/'
    complete_urls = [url_prefix + urlBody + urlExtend for urlBody, urlExtend in zip(urlBodies, urlExtends)]
    return address, complete_urls


def get_info(hotelId):
    global user_agents
    random_proxy = select_random_proxy(proxy_list)
    ip, port, username, password = random_proxy.split(':')
    proxy_url = f"http://{username}:{password}@{ip}:{port}"
    proxies = {
        'http': proxy_url,
        'https': proxy_url
    }
    burp0_url = "https://m.ctrip.com:443/webapp/hotels/sellingpoint?hotelid="+str(hotelId)+"&checkin=2023-10-11&checkout=2023-10-12&fromminiapp=weixin&allianceid=262684&sid=711465&sourceid=55552689&_cwxobj=%7B%22cid%22%3A%2252271115296426477628%22%2C%22appid%22%3A%22wx0e6ed4f51db9d078%22%2C%22mpopenid%22%3A%22b9787617-697b-4fe6-9942-f5fddd4b8063%22%2C%22mpunionid%22%3A%22oHkqHt8Zg6-uBtF-Y2UBUVVk1MlM%22%2C%22allianceid%22%3A%22262684%22%2C%22sid%22%3A%22711465%22%2C%22ouid%22%3A%22%22%2C%22sourceid%22%3A%2255552689%22%2C%22exmktID%22%3A%22%7B%5C%22openid%5C%22%3A%5C%22b9787617-697b-4fe6-9942-f5fddd4b8063%5C%22%2C%5C%22unionid%5C%22%3A%5C%22oHkqHt8Zg6-uBtF-Y2UBUVVk1MlM%5C%22%2C%5C%22channelUpdateTime%5C%22%3A%5C%221697011325536%5C%22%2C%5C%22serverFrom%5C%22%3A%5C%22WAP%2FWECHATAPP%5C%22%2C%5C%22innersid%5C%22%3A%5C%22%5C%22%2C%5C%22innerouid%5C%22%3A%5C%22%5C%22%2C%5C%22pushcode%5C%22%3A%5C%22%5C%22%2C%5C%22txCpsId%5C%22%3A%5C%22%5C%22%2C%5C%22amsPid%5C%22%3A%5C%22%5C%22%2C%5C%22gdt_vid%5C%22%3A%5C%22%5C%22%7D%22%2C%22scene%22%3A1007%2C%22personalRecommendSwitch%22%3Atrue%2C%22localRecommendSwitch%22%3Atrue%2C%22marketSwitch%22%3Atrue%2C%22pLen%22%3A3%7D&_obt=1697012693781"
    burp0_cookies = {"GUID": "52271115296426477628", "nfes_isSupportWebP": "1", "nfes_isSupportWebP": "1", "UBT_VID": "1697008100680.1e1cmmiBfNEU", "librauuid": "", "_RF1": "111.22.74.62", "_RSG": "2Ha6Cu221L4XAzI9a6MT2B", "_RDG": "282a01bdfbf8df2e6720e1b8de10ac1c86", "_RGUID": "4541082e-0c1a-4442-b09d-5869094bd2a1", "login_type": "0", "login_uid": "B34F49EE0A13E8463539186D23B00A8B", "DUID": "u=B34F49EE0A13E8463539186D23B00A8B&v=0", "IsNonUser": "F", "AHeadUserInfo": "VipGrade=10&VipGradeName=%BB%C6%BD%F0%B9%F3%B1%F6&UserName=&NoReadMessageCount=0", "cticket": "8A38C2D39415D5002658D2670FDD4F8D369AEE5E25A10B5F129505F540395603", "_resDomain": "https%3A%2F%2Fbd-s.tripcdn.cn", "_pd": "%7B%22_o%22%3A4%2C%22s%22%3A13%2C%22_s%22%3A0%7D", "MKT_Pagesource": "H5", "Union": "OUID=&AllianceID=262684&SID=711465&SourceID=55552689&AppID=wx0e6ed4f51db9d078&OpenID=b9787617-697b-4fe6-9942-f5fddd4b8063&exmktID={\"openid\":\"b9787617-697b-4fe6-9942-f5fddd4b8063\",\"unionid\":\"oHkqHt8Zg6-uBtF-Y2UBUVVk1MlM\",\"channelUpdateTime\":\"1697011325536\",\"serverFrom\":\"WAP/WECHATAPP\",\"innersid\":\"\",\"innerouid\":\"\",\"pushcode\":\"\",\"txCpsId\":\"\",\"amsPid\":\"\",\"gdt_vid\":\"\"}&createtime=1697012402&Expires=1697617201862", "MKT_OrderClick": "ASID=262684711465&AID=262684&CSID=711465&OUID=&CT=1697012401865&CURL=https%3A%2F%2Fm.ctrip.com%2Fwebapp%2Fservicechatv2%2F%3FisHideNavBar%3DYES%26isFreeLogin%3D0%26platform%3Dwechat%26appId%3Dwx0e6ed4f51db9d078%26sceneCode%3D2%26channel%3DEBK%26bizType%3D1356%26isPreSale%3D1%26pageCode%3D10320654891%26thirdPartytoken%3Dae1f21df-6e4b-4473-bcf3-8f19a78ac594%26source%3Dminipro_app%26orderInfo%3D%257B%2522amount%2522%253A%2522%2522%252C%2522bu%2522%253A%2522EBK%2522%252C%2522cid%2522%253A%25220%2522%252C%2522ctype%2522%253A%2522%2522%252C%2522currency%2522%253A%2522%2522%252C%2522supplierId%2522%253A1632483%252C%2522supplierName%2522%253A%2522%25E6%2598%259F%25E8%25BE%25B0%25E9%2585%2592%25E5%25BA%2597(%25E6%25A0%25AA%25E6%25B4%25B2%25E6%25B9%2596%25E5%258D%2597%25E5%25B7%25A5%25E4%25B8%259A%25E5%25A4%25A7%25E5%25AD%25A6%25E5%25BA%2597)%2522%252C%2522title%2522%253A%2522%25E6%2598%259F%25E8%25BE%25B0%25E9%2585%2592%25E5%25BA%2597(%25E6%25A0%25AA%25E6%25B4%25B2%25E6%25B9%2596%25E5%258D%2597%25E5%25B7%25A5%25E4%25B8%259A%25E5%25A4%25A7%25E5%25AD%25A6%25E5%25BA%2597)%2522%257D%26q%3DeyJtaW5lIjowLCJ3ZW1jIwoxfQ%3D%3DTW%26mktopenid%3Db9787617-697b-4fe6-9942-f5fddd4b8063%26fromminiapp%3Dweixin%26allianceid%3D262684%26sid%3D711465%26sourceid%3D55552689%26_cwxobj%3D%257B%2522cid%2522%253A%252252271115296426477628%2522%252C%2522appid%2522%253A%2522wx0e6ed4f51db9d078%2522%252C%2522mpopenid%2522%253A%2522b9787617-697b-4fe6-9942-f5fddd4b8063%2522%252C%2522mpunionid%2522%253A%2522oHkqHt8Zg6-uBtF-Y2UBUVVk1MlM%2522%252C%2522allianceid%2522%253A%2522262684%2522%252C%2522sid%2522%253A%2522711465%2522%252C%2522ouid%2522%253A%2522%2522%252C%2522sourceid%2522%253A%252255552689%2522%252C%2522exmktID%2522%253A%2522%257B%255C%2522openid%255C%2522%253A%255C%2522b9787617-697b-4fe6-9942-f5fddd4b8063%255C%2522%252C%255C%2522unionid%255C%2522%253A%255C%2522oHkqHt8Zg6-uBtF-Y2UBUVVk1MlM%255C%2522%252C%255C%2522channelUpdateTime%255C%2522%253A%255C%25221697011325536%255C%2522%252C%255C%2522serverFrom%255C%2522%253A%255C%2522WAP%252FWECHATAPP%255C%2522%252C%255C%2522innersid%255C%2522%253A%255C%2522%255C%2522%252C%255C%2522innerouid%255C%2522%253A%255C%2522%255C%2522%252C%255C%2522pushcode%255C%2522%253A%255C%2522%255C%2522%252C%255C%2522txCpsId%255C%2522%253A%255C%2522%255C%2522%252C%255C%2522amsPid%255C%2522%253A%255C%2522%255C%2522%252C%255C%2522gdt_vid%255C%2522%253A%255C%2522%255C%2522%257D%2522%252C%2522scene%2522%253A1007%252C%2522personalRecommendSwitch%2522%253Atrue%252C%2522localRecommendSwitch%2522%253Atrue%252C%2522marketSwitch%2522%253Atrue%252C%2522pLen%2522%253A3%257D&VAL={\"h5_vid\":\"1697008100680.1e1cmmiBfNEU\"}", "_bfa": "1.1697008100680.1e1cmmiBfNEU.1.1697010109935.1697012401870.1.20.10650084702", "_ubtstatus": "%7B%22vid%22%3A%221697008100680.1e1cmmiBfNEU%22%2C%22sid%22%3A1%2C%22pvid%22%3A20%2C%22pid%22%3A10650084702%7D"}
    burp0_headers = {"Upgrade-Insecure-Requests": "1", "User-Agent":random.choice(user_agents), "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9", "Sec-Fetch-Site": "none", "Sec-Fetch-Mode": "navigate", "Sec-Fetch-User": "?1", "Sec-Fetch-Dest": "document", "Sec-Ch-Ua": "\"Chromium\";v=\"107\", \"Not=A?Brand\";v=\"24\"", "Sec-Ch-Ua-Mobile": "?0", "Sec-Ch-Ua-Platform": "\"Windows\"", "Accept-Encoding": "gzip, deflate", "Accept-Language": "zh-CN,zh;q=0.9"}
    time.sleep(5)
    res = requests.get(burp0_url, headers=burp0_headers, cookies=burp0_cookies, proxies=proxies)
    time.sleep(5)
    print(f"酒店电话，开业时间等使用{proxies}爬取")
    print("这次用了" + burp0_headers["User-Agent"] + "请求头")
    soup = BeautifulSoup(res.text, "html.parser")
    script_tag = soup.find("script", {"id": "__NEXT_DATA__"})
    json_data = json.loads(script_tag.string)
    initialState = json_data["props"]["pageProps"]["initialState"]
    hotelName = initialState.get("hotelEnName", "")
    tels = [tel.get("calTel", "") for tel in initialState.get("contactInfo", {}).get("telInfoList", [])]
    email = initialState.get("contactInfo", {}).get("email", "")
    profile = initialState.get("sectionList", [{}])[1].get("desc", "")
    openTime = next((info.get("text", "") for info in initialState.get("baseInfo", []) if info.get("name") == "开业时间"),
                    "")
    renewTime = next((info.get("text", "") for info in initialState.get("baseInfo", []) if info.get("name") == "装修时间"),
                     "")
    roomNums = next((info.get("text", "") for info in initialState.get("baseInfo", []) if info.get("name") == "房间数量"),
                    "")
    return hotelName, tels, openTime, roomNums, renewTime, email, profile


def saveE(name, data, headers):
    # 创建一个新的文件夹
    folder_name = name  # 文件夹名称与传入的 name 参数一致
    os.makedirs(folder_name, exist_ok=True)
    # 调用 download_and_update_images 处理第一个列
    excel_file_path_1 = save_excel_file(folder_name, f"{name}酒店信息", data[0], headers[0])
    excel_file_path_2 = save_excel_file(folder_name, f"{name}酒店评论信息", data[1], headers[1])
    print("这是在调用saveE")
    return excel_file_path_1, excel_file_path_2

import string

def clean_text(text):
    # 创建一个可接受字符的白名单
    whitelist = string.printable
    # 使用白名单字符来过滤文本
    cleaned_text = ''.join(c for c in text if c in whitelist)
    return cleaned_text

def save_excel_file(folder_name, excel_name, data, header):
    # 创建一个新的Excel文件
    workbook = Workbook()
    # 创建第一个表
    sheet1 = workbook.active
    sheet1.title = "表1"
    # 设置第一个表的表头标题栏
    sheet1.append(header)
    y = 2
    for v in data:
        v = [clean_text(value) for value in v]  # 清理数据
        x = 1
        for vv in v:
            sheet1.cell(row=y, column=x, value=vv)
            x += 1
        y += 1
    # for v in data:
    #     x = 1
    #     for vv in v:
    #         sheet1.cell(row=y, column=x, value=vv)
    #         x += 1
    #     y += 1
    # 保存Excel文件
    excel_file_path = os.path.join(folder_name, f"{excel_name}.xlsx")
    workbook.save(filename=excel_file_path)
    print("这是在调用save_excel_file")
    return excel_file_path


def read_hotel_ids_from_file(file_path):
    hotel_ids = []
    with open(file_path, 'r') as file:
        for line in file:
            hotel_id = int(line.strip())
            hotel_ids.append(hotel_id)
    return hotel_ids


def process_hotel_data(cityid):
    # 生成和保存酒店ID到txt文件
    txt_file_name = save_hotel_ids_to_txt(cityid)
    city_name = os.path.splitext(txt_file_name)[0]
    # 读取酒店ID
    hotel_ids = read_hotel_ids_from_file(txt_file_name)
    # 处理酒店信息

    headers = [
        ["酒店名称", "地址", "电话","开业时间","装修时间","客房数","邮箱","简介","页面网址","酒店图片"],
        ["酒店名称", "评论内容", "图片链接"]
    ]
    data1 = []
    data2 = []
    error_hotel_ids = []  # 用于存储报错的酒店 ID
    for hotelId in hotel_ids:
        try:
            # 可能会引发异常的代码
            errorHotelId = hotelId
            hotel_url = f"https: // hotels.ctrip.com / hotels / detail /?hotelId = {hotelId}"
            hotelName, tels, openTime, renewTime, roomNums, email, profile = get_info(hotelId)
            print(hotelName, tels, openTime, renewTime, roomNums)
            address, pics = get_addressApic(hotelId)
            print(hotelName, len(pics), "张图片已经获取")
            comments_list, comments, comments_image_url_list = get200comments(hotelId)
            print(len(comments_list), "条点评已经获取")
            num_comments = len(comments_list)
            hotel_info1 = [hotelName, address, "\n".join(tels), openTime, renewTime, roomNums, email, profile,
                           hotel_url, "\n".join(pics)]
            hotel_info2 = [hotelName]
            for i in range(num_comments):
                comment = comments_list[i] if i < len(comments_list) else ""
                image_urls = "\n".join(comments_image_url_list[i]) if i < len(comments_image_url_list) else ""
                # 合并酒店信息和评论信息
                row_data = hotel_info2 + [comment, image_urls]
                # 将一行数据写入工作表
                data2.append(row_data)
            data1.append(hotel_info1)
        except Exception as e:
            print(f"发生了一个异常：{str(e)}")
            print(f"已经紧急保存excel文件，报错的酒店id: {errorHotelId}")
            error_hotel_ids.append(str(errorHotelId))
            if data1:
                saveE(city_name, [data1, data2], headers)
            print("这是在调用process")
        saveE(city_name, [data1, data2], headers)
        print("这是在调用process")

    # 将报错的酒店 ID 写入到 txt 文件
    error_file_path = os.path.join(city_name, f"{city_name}_error.txt")
    with open(error_file_path, 'w') as error_file:
        error_file.write("\n".join(error_hotel_ids))
    print("已将报错的酒店 ID 写入到文件:", error_file_path)
    print("这是在调用process")

